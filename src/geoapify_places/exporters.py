"""Helpers for exporting sweep JSON output to Excel files."""

from __future__ import annotations

import json
import csv
from pathlib import Path
from typing import Iterable, List, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

Record = Mapping[str, object]


def flatten_records(records: Sequence[Record]) -> tuple[list[str], list[list[object]]]:
    """
    Flatten a list of JSON-friendly dictionaries into table rows.

    Lists are converted to comma-separated strings, while nested dicts
    are serialized as JSON blobs to keep the output Excel-friendly.
    """

    headers: List[str] = []
    rows: List[List[object]] = []

    for record in records:
        for key in record.keys():
            if key not in headers:
                headers.append(key)

    for record in records:
        row: List[object] = []
        for header in headers:
            value = record.get(header)
            row.append(_coerce(value))
        rows.append(row)

    return headers, rows


def export_to_excel(
    records: Sequence[Record],
    output_path: str | Path,
    *,
    sheet_name: str = "Businesses",
) -> Path:
    """Write JSON-like records to an Excel workbook."""

    if not records:
        raise ValueError("No records supplied for export")

    headers, rows = flatten_records(records)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    _auto_size_columns(worksheet, headers, rows)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_to_csv(
    records: Sequence[Record],
    output_path: str | Path,
) -> Path:
    """Write JSON-like records to a CSV file."""

    if not records:
        raise ValueError("No records supplied for export")

    headers, rows = flatten_records(records)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


def _coerce(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    if isinstance(value, Mapping):
        return json.dumps(value, ensure_ascii=False)
    return value


def _auto_size_columns(worksheet, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    widths = [len(str(header)) for header in headers]
    for row in rows:
        for idx, cell in enumerate(row):
            widths[idx] = max(widths[idx], len(str(cell)))

    for idx, width in enumerate(widths, start=1):
        column_letter = get_column_letter(idx)
        worksheet.column_dimensions[column_letter].width = min(width + 2, 60)


__all__ = ["export_to_excel", "export_to_csv", "flatten_records"]
